import openpyxl
from tkinter import messagebox
import pandas as pd
from datetime import date
import os 

class DataModification:
    def __init__(self):
        self.weekly_row = 0
        self.current_directory = os.path.join(os.getcwd(), "Desktop/StockSavvy/_internal")
        print("data file: "+ self.current_directory)
    
    def create_user(self):
        username = self.textvariables["username"].get()
        password = self.textvariables["password"].get()
        re_password = self.textvariables["re_password"].get()
        user_authentication = self.authenticate(username)
        if user_authentication[0]:
            signup = messagebox.askokcancel(message="User exists \n please login.")
            if signup:
                self.main_page()
            else:
                pass
        
        else:
            if password == re_password:
                if len(password) >= 8 and len(password) <= 16:
                    path = os.path.join(self.current_directory ,"Data/Authentication.xlsx")
                    wb = openpyxl.load_workbook(path)
                    sheet = wb.active
                    sheet.append((username, password))
                    signup = messagebox.askokcancel(message="Account created \n Please Login.")
                    wb.save(path)
                    if signup:
                        self.main_page()
                    else:
                        pass   
                else:
                    messagebox.showerror(message="Password must be 8 to 16 characters")
            else:
                messagebox.showerror(message="Password does not match")

    def change_password(self):
        username = self.textvariables["username"].get()
        password = self.textvariables["password"].get()
        re_password = self.textvariables["re_password"].get()
        user_authentication = self.authenticate(username)
        if user_authentication[0]:
            if password == re_password:
                if len(password) >= 8 and len(password) <= 16:
                    path = os.path.join(self.current_directory ,"Data/Authentication.xlsx")
                    wb = openpyxl.load_workbook(path)
                    sheet = wb.active
                    sheet["B" + str(user_authentication[2])] = password
                    signup = messagebox.askokcancel(message="Password has been changed \n Please Login.")
                    wb.save(path)
                    if signup:
                        self.main_page()
                    else:
                        pass   
                else:
                    messagebox.showerror(message="Password must be 8 to 16 characters")
            else:
                messagebox.showerror(message="Password does not match")

        else:
            signup = messagebox.askokcancel(message="User does not exists \n Please Sign Up.")
            if signup:
                self.signup_page()
            else:
                pass
    
    def authenticate(self, username):
        exists = False
        credentials = []
        path = os.path.join(self.current_directory ,"Data/Authentication.xlsx")
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        max_row = sheet.max_row
        for i in range(2, max_row+1):
            if sheet.cell(row=i, column=1).value == username:
                exists = True
                credentials.append(username)
                credentials.append(sheet.cell(row=i, column=2).value)
                break
        return (exists, credentials, i)

    def check_credentials(self):
        user_authentication = self.authenticate(self.textvariables["username"].get())
        if user_authentication[0]:
            if self.textvariables["username"].get() == user_authentication[1][0] and self.textvariables["password"].get() == user_authentication[1][1]:
                self.buttons["verify"].configure(text="Continue", command=self.stock_page)
                
            else:
                messagebox.showwarning("Error", "wrong username or password")
                # Clear TextVariables
                self.textvariables["username"].set("")
                self.textvariables["password"].set("")
        else:
            messagebox.showinfo(message="User doesn't exists \n please Sign Up.")
        

    def weeklyUpdate(self, stock):
        values = []
        path = os.path.join(self.current_directory ,"Data/Weekly Update.xlsx")
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        for i in range(2, max_row+1):
            if sheet.cell(row=i, column=1).value == stock:
                self.weekly_row = i
                break
        if self.weekly_row == 0:
            print("error")
        else:
            for j in range(2, max_column+1):
                values.append(sheet.cell(row=self.weekly_row, column=j).value)

        return values
    
    def ModifyGraph(self):
        file_path = self.textvariables['graph'].get()  
        self.data = pd.read_csv(file_path)
        self.data['date'] = pd.to_datetime(self.data['date'])
        self.data.set_index('date', inplace=True)
        self.plot()
    
    def ClosingValue(self, stock):
        close = pd.read_csv(os.path.join(self.current_directory ,f'Data/{stock}/Stock_ML_CLOSE.csv'))
        week_close = close['Expected'].tail(5).to_list()
        today = date.today().weekday()
        
        if today+1 == 5 or today+1 == 6:
            return week_close[4]
        elif today+1 == 7:
            return week_close[0]
        else:
            return week_close[today+1]

    
        

        